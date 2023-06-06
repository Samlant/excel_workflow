from openpyxl import Workbook, load_workbook


class sheet_creator:
    def __init__(self, month: str) -> None:
        self.month = month
        self.header_list = [
            "X",
            "*",
            "ME",
            "Client",
            "Start",
            "Need",
            "Year",
            "Vessel",
            "Markets",
            "CH",
            "MK",
            "AI",
            "AM",
            "PG",
            "SW",
            "KM",
            "CP",
            "YI",
            "IN",
            "TV",
            "Call",
            "Rec",
            "Mkt Bnd",
            "Last Update",
            "Referral",
        ]
        self.create_sheet()

    def save(self) -> None:
        self.wb.save("test.xlsx")

    def create_sheet(self) -> None:
        self.wb = load_workbook("test.xlsx")
        self.ws = wb.create_sheet(self.month)
        self._create_header()

    def _create_header(self) -> None:
        self.ws.append(self.header_list)

    # ws["U1"] = "Call"
    # ws["V1"] = "Rec"
    # ws["W1"] = "Mkt Bnd"
    # ws["X1"] = "Last Update"
    # ws["Y1"] = "Referral"

    # ws.title = "June"
